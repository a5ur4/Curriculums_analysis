import json
import os
import pytesseract
import cv2
import openpyxl
import re
import logging
import requests
from pdf2image import convert_from_path

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

folder_path = r"C:\Users\pedro\Downloads\curriculums"
path_tesseract = r"C:\Users\pedro\AppData\Local\Programs\Tesseract-OCR"
poppler_path = r"C:\poppler\Library\bin"
save_path = r"C:\Users\pedro\Downloads\curriculums\sheets"
txt_save_path = r"C:\Users\pedro\Downloads\curriculums\txt_files"
pytesseract.pytesseract.tesseract_cmd = os.path.join(path_tesseract, "tesseract.exe")

keywords = ["Java", "Spring", "Python", "Django", "JavaScript", "React", "Node", 
            "SQL", "NoSQL", "MongoDB", "PostgreSQL", "MySQL", "HTML", "CSS", "Bootstrap"]

results = []
experiences = []

def convert_to_image(file_path, poppler_path, save_path):
    try:
        images = convert_from_path(file_path, poppler_path=poppler_path)
        image_paths = []
        for i, image in enumerate(images):
            image_path = os.path.join(save_path, f"{os.path.basename(file_path)}_page_{i}.png")
            image.save(image_path, "PNG")
            image_paths.append(image_path)
        return image_paths
    except Exception as e:
        logging.error(f"Error converting PDF to images: {file_path} - {e}")
        return []

def extract_text(image_path):
    try:
        text = pytesseract.image_to_string(cv2.imread(image_path), lang="eng")
        return text
    except Exception as e:
        logging.error(f"Error extracting text from image: {image_path} - {e}")
        return ""

def save_to_json(data, json_save_path, filename):
    try:
        json_file_path = os.path.join(json_save_path, f"{os.path.splitext(filename)[0]}.json")
        with open(json_file_path, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        logging.info(f"Data saved to JSON: {json_file_path}")
    except Exception as e:
        logging.error(f"Error saving data to JSON: {filename} - {e}")

def extract_name_from_text(text):
    lines = text.splitlines()
    for line in lines[:2]:
        for name in line.split():
            if name.lower() in line.lower():
                return name
    return "Name not found"

def extract_contact(text):
    phone_pattern = r'\(?\d{2}\)?\s?\d{4,5}-?\d{4}'
    email_pattern = r'[a-zA-Z0-9._%+-]+\s?@\s?[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'

    phones = re.findall(phone_pattern, text)
    emails = re.findall(email_pattern, text)
    
    return {
        "phones": phones if phones else ["Phone not found"],
        "emails": emails if emails else ["Email not found"]
    }

def extract_experience(text):
    experiences.clear()
    date_pattern = r'.*(20[0-2][0-9]|202[0-5]).*'

    lines = text.split('\n')
    for line in lines:
        if re.search(date_pattern, line):
            experiences.append(line.strip())
    
    return list(set(experiences))

def approved(found_keywords):
    return len(found_keywords) > 5

def create_sheet(results):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # Add new column for LLM Summary
        headers = ["File", "Name", "Phones", "Emails", "Approved", "Experiences", "LLM Summary"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        logging.info("Inserting data into spreadsheet...")
        for i, result in enumerate(results, start=2):
            if not isinstance(result, dict):
                logging.error(f"Invalid result format: {result}")
                continue

            ws[f"A{i}"] = result.get("File", "Unknown")
            ws[f"B{i}"] = result.get("Name", "Name not found")
            ws[f"C{i}"] = ", ".join(result.get("Phones", ["Phone not found"]))
            ws[f"D{i}"] = ", ".join(result.get("Emails", ["Email not found"]))
            ws[f"E{i}"] = "Yes" if result.get("Approved", False) else "No"
            ws[f"F{i}"] = "\n".join(result.get("Experiences", []))
            ws[f"G{i}"] = result.get("LLM Summary", "No summary available")

        logging.info("Adjusting column widths...")
        for col in ws.columns:
            max_length = 0
            column = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        if not os.path.exists(save_path):
            os.makedirs(save_path)

        output_file = os.path.join(save_path, "results.xlsx")
        wb.save(output_file)
        logging.info(f"Results spreadsheet successfully created at {output_file}!")

    except Exception as e:
        logging.error(f"Error creating spreadsheet: {e}")

def send_to_llm(data):
    url = "http://localhost:11434/api/generate"
    headers = {"Content-Type": "application/json"}
    prompt_message = '''Generate a candidate summary for a software developer position. Structure the response as:

    [Core Competencies]
    - 3-5 key technical strengths

    [Key Experience Highlights]
    - 2-3 notable positions/achievements with metrics if available

    [Technical Skills]
    - Categorize skills (e.g., "Languages: Java, Python | Frameworks: Spring, React")

    [Education]
    - Degree and institution (if mentioned)

    Keep responses:
    1. Succinct (max 300 characters)
    2. Use bullet points and abbreviations
    3. Focus on software development requirements
    4. Exclude personal pronouns/narrative

    Resume Content: ''' # Add your prompt message here
    payload = {
        "model": "llama3.2", # Change to your model, i'm using llama3.2 with 3b parameters, so if you want a better result, change to a bigger model
        "prompt": f"{prompt_message}\n\n{json.dumps(data, ensure_ascii=False)}",
        "stream": False
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        return response.json().get("response", "No summary generated")
    except Exception as e:
        logging.error(f"Error sending to LLM: {e}")
        return None

# Create necessary directories
os.makedirs(save_path, exist_ok=True)
os.makedirs(txt_save_path, exist_ok=True)

# Process PDF files
for filename in os.listdir(folder_path):
    if filename.endswith(".pdf"):
        file_path = os.path.join(folder_path, filename)
        logging.info(f"Processing file: {filename}")
        
        image_paths = convert_to_image(file_path, poppler_path, save_path)
        all_text = ""

        for image_path in image_paths:
            text = extract_text(image_path)
            save_to_json(text, txt_save_path, filename)
            all_text += text
            os.remove(image_path)
        
        txt_file_path = os.path.join(txt_save_path, f"{os.path.splitext(filename)[0]}.txt")
        with open(txt_file_path, "w", encoding="utf-8") as txt_file:
            txt_file.write(all_text)

        # Extract information
        name = extract_name_from_text(all_text)
        contact_info = extract_contact(all_text)
        experience_info = extract_experience(all_text)
        found_keywords = [kw for kw in keywords if kw.lower() in all_text.lower()]
        is_approved = approved(found_keywords)

        # Create result entry
        result = {
            "File": filename,
            "Name": name,
            "Phones": contact_info["phones"],
            "Emails": contact_info["emails"],
            "Approved": is_approved,
            "Experiences": experience_info
        }

        # Get LLM summary and add to result
        llm_summary = send_to_llm(result)
        if llm_summary:
            result["LLM Summary"] = llm_summary
        
        results.append(result)

# Create final spreadsheet
create_sheet(results)

# Made by: Pedro Bastos - a5ur4